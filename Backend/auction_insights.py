import os
from dotenv import load_dotenv
from supabase import create_client
from collections import defaultdict
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from googleapiclient.http import MediaFileUpload
from drive_utils import get_drive_service

# ==========================================
# 1. SETUP & ENVIRONMENT
# ==========================================
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
load_dotenv(os.path.join(BASE_DIR, ".env"))

url = os.environ.get("SUPABASE_URL")
key = os.environ.get("SUPABASE_KEY")


def _normalize_customer_id(value) -> str:
    raw = str(value or "").strip()
    digits = "".join(ch for ch in raw if ch.isdigit())
    return digits or raw


def _customer_id_candidates(value) -> list[str]:
    raw = str(value or "").strip()
    normalized = _normalize_customer_id(raw)
    out = []
    for candidate in (raw, normalized):
        if candidate and candidate not in out:
            out.append(candidate)
    return out

def get_supabase():
    if not url or not key:
        raise ValueError("Missing Supabase URL or Key in .env file.")
    return create_client(url, key)


# ==========================================
# UPLOAD TO GOOGLE DRIVE
# ==========================================
def upload_excel_to_drive(local_filename: str, drive_folder_id: str):
    drive = get_drive_service()
    print(f"Uploading '{local_filename}' to Google Drive...")

    try:
        file_metadata = {
            'name': os.path.basename(local_filename),
            'parents': [drive_folder_id]
        }

        media = MediaFileUpload(
            local_filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            resumable=False,
        )
        uploaded_file = drive.files().create(
            body=file_metadata,
            media_body=media,
            fields='id,webViewLink,webContentLink,name,parents',
            supportsAllDrives=True
        ).execute()

        file_id = uploaded_file.get('id')
        file_link = uploaded_file.get('webViewLink') or uploaded_file.get('webContentLink')
        if not file_link and file_id:
            # Server/shared-drive responses may omit webViewLink; build deterministic URL.
            file_link = f"https://drive.google.com/file/d/{file_id}/view"

        if file_id:
            verify = drive.files().get(
                fileId=file_id,
                fields='id,parents,webViewLink,webContentLink',
                supportsAllDrives=True
            ).execute()
            parents = verify.get('parents') or []
            if drive_folder_id not in parents:
                print(f"⚠️ Uploaded file parent mismatch. Expected {drive_folder_id}, got {parents}")
            file_link = verify.get('webViewLink') or verify.get('webContentLink') or file_link
            print(f"Uploaded file ID: {file_id}")
        print(f"✅ Successfully uploaded! View it here: {file_link}")

        if os.path.exists(local_filename):
            os.remove(local_filename)
            print("🧹 Cleaned up temporary local file.")

        return file_link

    except Exception as e:
        print(f"❌ Failed to upload to Google Drive: {e}")
        return None


# ==========================================
# 2. CONVERT "January 2026" → start & end dates
# ==========================================
def month_text_to_date_range(month_text: str):
    try:
        dt = datetime.strptime(str(month_text).strip(), "%B %Y")
        start_date = dt.strftime("%Y-%m-%d")
        import calendar
        last_day = calendar.monthrange(dt.year, dt.month)[1]
        end_date = dt.strftime(f"%Y-%m-{last_day:02d}")
        return start_date, end_date
    except Exception as e:
        print(f"⚠️ Could not parse month '{month_text}': {e}")
        return None, None


# ==========================================
# 3. GET LATEST REPORT DATE FOR CUSTOMER
# ==========================================
def get_latest_report_date(customer_id: str):
    sb = get_supabase()
    for candidate in _customer_id_candidates(customer_id):
        response = sb.table("google_ads_auction_master") \
                     .select("report_date") \
                     .eq("customer_id", candidate) \
                     .order("report_date", desc=True) \
                     .limit(1) \
                     .execute()
        if response.data:
            latest = response.data[0]["report_date"]
            print(f"📅 Latest report_date found: {latest} (customer_id={candidate})")
            return latest
    return None


# ==========================================
# 4. DATA FETCH & AGGREGATION LOGIC
# ==========================================
def get_auction_insights_data(
    customer_id: str,
    report_month: str = None,
    start_date: str = None,
    end_date: str = None,
):
    sb = get_supabase()

    if start_date and end_date:
        display_label = report_month or f"{start_date} → {end_date}"
    elif report_month:
        start_date, end_date = month_text_to_date_range(report_month)
        if not start_date:
            print(f"❌ Invalid month format: {report_month}. Use 'January 2026'.")
            return {}
        display_label = report_month
    else:
        latest = get_latest_report_date(customer_id)
        if not latest:
            print(f"❌ No data found for Customer ID: {customer_id}")
            return {}
        start_date = end_date = latest
        display_label = latest

    print(f"Fetching DB -> Customer: {customer_id} | Range: {start_date} → {end_date}")

    try:
        raw_data = []
        matched_customer_id = None
        for candidate in _customer_id_candidates(customer_id):
            response = sb.table("google_ads_auction_master") \
                         .select("account_name, campaign, domain, impression_share, position_above_rate, top_of_page_rate, absolute_top_of_page_rate, report_date") \
                         .eq("customer_id", candidate) \
                         .gte("report_date", start_date) \
                         .lte("report_date", end_date) \
                         .execute()
            raw_data = response.data or []
            if raw_data:
                matched_customer_id = candidate
                break

        if not raw_data:
            print(f"❌ No rows found for {customer_id} between {start_date} → {end_date}")
            return {}
        if matched_customer_id:
            print(f"✅ Auction rows matched using customer_id={matched_customer_id}")

        grouped = defaultdict(lambda: {
            "account_name": "",
            "domains": defaultdict(lambda: {
                "impression_share_sum": 0.0,
                "position_above_rate_sum": 0.0,
                "top_of_page_rate_sum": 0.0,
                "absolute_top_of_page_rate_sum": 0.0,
                "count": 0
            })
        })

        for row in raw_data:
            campaign = row.get("campaign") or "Unknown Campaign"
            domain = row.get("domain") or "Unknown"
            account = row.get("account_name") or "Unknown Account"

            grouped[campaign]["account_name"] = account
            grouped[campaign]["domains"][domain]["impression_share_sum"] += float(row.get("impression_share") or 0.0)
            grouped[campaign]["domains"][domain]["position_above_rate_sum"] += float(row.get("position_above_rate") or 0.0)
            grouped[campaign]["domains"][domain]["top_of_page_rate_sum"] += float(row.get("top_of_page_rate") or 0.0)
            grouped[campaign]["domains"][domain]["absolute_top_of_page_rate_sum"] += float(row.get("absolute_top_of_page_rate") or 0.0)
            grouped[campaign]["domains"][domain]["count"] += 1

        formatted = {}

        for campaign, data in grouped.items():
            competitors = []
            for domain, metrics in data["domains"].items():
                count = metrics["count"]
                
                # FIX: Removed rounding to keep the exact decimal values (e.g., 0.3831)
                competitors.append({
                    "domain": domain,
                    "impression_share": metrics["impression_share_sum"] / count,
                    "position_above_rate": metrics["position_above_rate_sum"] / count,
                    "top_of_page_rate": metrics["top_of_page_rate_sum"] / count,
                    "absolute_top_of_page_rate": metrics["absolute_top_of_page_rate_sum"] / count,
                })

            competitors.sort(key=lambda x: x["impression_share"], reverse=True)

            formatted[campaign] = {
                "campaign": campaign,
                "account_name": data["account_name"],
                "report_month": display_label,
                "date_range": f"{start_date} → {end_date}",
                "competitors": competitors
            }

        return formatted

    except Exception as e:
        print(f"❌ Database error: {e}")
        return {}


# ==========================================
# 5. PRINT HELPER
# ==========================================
def print_auction_table(data: dict):
    if not data:
        print("No data to display.")
        return

    for campaign, campaign_data in data.items():
        print(f"\n{'='*95}")
        print(f"  Campaign     : {campaign}")
        print(f"  Account      : {campaign_data['account_name']}")
        print(f"  Month        : {campaign_data['report_month']}")
        print(f"  Date Range   : {campaign_data['date_range']}")
        print(f"{'='*95}")
        print(f"  {'Display URL Domain':<30} {'Impr. Share':>12} {'Pos. Above':>12} {'Top of Page':>12} {'Abs. Top':>12}")
        print(f"  {'-'*30} {'-'*12} {'-'*12} {'-'*12} {'-'*12}")

        for c in campaign_data["competitors"]:
            # Multiply by 100 here just for terminal display formatting
            print(f"  {c['domain']:<30} {c['impression_share']*100:>11.2f}% {c['position_above_rate']*100:>11.2f}% {c['top_of_page_rate']*100:>11.2f}% {c['absolute_top_of_page_rate']*100:>11.2f}%")

    print(f"\n{'='*95}\n")


# ==========================================
# GENERATE STANDALONE EXCEL FILE (.XLSX)
# ==========================================
def _safe_filename_part(text: str) -> str:
    import re
    cleaned = re.sub(r"[^A-Za-z0-9]+", "_", str(text or "")).strip("_")
    return cleaned or "Account"


def generate_auction_xlsx(
    customer_id: str,
    report_month: str = None,
    dealer_name: str = None,
    start_date: str = None,
    end_date: str = None,
):
    data = get_auction_insights_data(
        customer_id,
        report_month=report_month,
        start_date=start_date,
        end_date=end_date,
    )

    if not data:
        print("❌ No data to export to Excel.")
        return None

    wb = Workbook()
    ws = wb.active
    ws.title = "Auction Insights"

    # Define the headers
    headers = [
        "Campaign",
        "Account Name",
        "Report Month",
        "Display URL Domain",
        "Impression Share",
        "Position Above Rate",
        "Top of Page Rate",
        "Abs. Top of Page Rate"
    ]
    ws.append(headers)

    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    center_align = Alignment(horizontal="center", vertical="center")
    left_align = Alignment(horizontal="left", vertical="center")

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align

    row_idx = 2
    you_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    for campaign, camp_data in data.items():
        for comp in camp_data["competitors"]:
            # Core data (A-D)
            ws.cell(row=row_idx, column=1, value=campaign).alignment = left_align
            ws.cell(row=row_idx, column=2, value=camp_data["account_name"]).alignment = left_align
            ws.cell(row=row_idx, column=3, value=camp_data["report_month"]).alignment = center_align
            ws.cell(row=row_idx, column=4, value=comp["domain"]).alignment = left_align

            # Clean and normalize float values (E-H)
            impression_val = comp["impression_share"] / 100 if comp["impression_share"] > 1 else comp["impression_share"]
            pos_above_val = comp["position_above_rate"] / 100 if comp["position_above_rate"] > 1 else comp["position_above_rate"]
            top_page_val = comp["top_of_page_rate"] / 100 if comp["top_of_page_rate"] > 1 else comp["top_of_page_rate"]
            abs_top_val = comp["absolute_top_of_page_rate"] / 100 if comp["absolute_top_of_page_rate"] > 1 else comp["absolute_top_of_page_rate"]

            # Impression Share (Col E)
            c5 = ws.cell(row=row_idx, column=5, value=impression_val)
            c5.number_format = '0.00%'
            c5.alignment = center_align

            # --- Start of user request fix: Column F (Position Above) ---
            is_you_row = str(comp["domain"]).strip().lower() == "you"
            
            # Use the exact logic from the request: "Make is empty ok on the position"
            # We explicitly check for 0.00% (0.0 float) and the "You" domain.
            if is_you_row and pos_above_val == 0.0:
                # If it's the "You" row and the value is zero, leave it blank (None value)
                c6 = ws.cell(row=row_idx, column=6, value=None)
            else:
                # For non-"You" rows, OR "You" rows with a non-zero value, pass the value as usual
                c6 = ws.cell(row=row_idx, column=6, value=pos_above_val)
                
            # Apply percentage formatting to all, which renders a 'None' value as empty.
            c6.number_format = '0.00%' 
            c6.alignment = center_align
            # --- End of fix ---

            # Top of Page (Col G)
            c7 = ws.cell(row=row_idx, column=7, value=top_page_val) 
            c7.number_format = '0.00%'
            c7.alignment = center_align

            # Abs Top (Col H)
            c8 = ws.cell(row=row_idx, column=8, value=abs_top_val)   
            c8.number_format = '0.00%'
            c8.alignment = center_align

            # Highlight the entire row yellow if domain is "You"
            if is_you_row:
                for col_num in range(1, 9):
                    ws.cell(row=row_idx, column=col_num).fill = you_fill

            row_idx += 1

    # Adjust Column Widths
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 30
    ws.column_dimensions['E'].width = 18
    ws.column_dimensions['F'].width = 20 
    ws.column_dimensions['G'].width = 18 
    ws.column_dimensions['H'].width = 22 

    # Generate a clean filename and save
    if start_date and end_date:
        safe_month = f"{start_date}_to_{end_date}"
    elif report_month:
        safe_month = report_month.replace(' ', '_')
    else:
        safe_month = 'Latest'

    name_for_file = dealer_name
    if not name_for_file:
        try:
            first_campaign = next(iter(data.values()), {})
            name_for_file = first_campaign.get("account_name")
        except Exception:
            name_for_file = None
    if not name_for_file:
        name_for_file = customer_id

    safe_name = _safe_filename_part(name_for_file)
    filename = f"Auction_Insights_{safe_name}_{safe_month}.xlsx"

    wb.save(filename)
    print(f"✅ Excel successfully generated and saved locally as: {filename}")

    return filename


# ==========================================
# 6. LOCAL TESTING BLOCK
# ==========================================
if __name__ == "__main__":
    print("\n=== LOCAL TESTING MODE ===")

    # WWheeler Digital Customer ID
    TEST_CUSTOMER_ID = "5691491477"
    TEST_REPORT_MONTH = "February 2026"
    
    # Wheels Digital Auction Insights folder
    TARGET_DRIVE_FOLDER_ID = "1pR1oWgzhA51YZm1c9MnZt3LULHLp_gAJ"

    generated_file = generate_auction_xlsx(TEST_CUSTOMER_ID, TEST_REPORT_MONTH)

    if generated_file:
        drive_link = upload_excel_to_drive(generated_file, TARGET_DRIVE_FOLDER_ID)
        if drive_link:
            print(f"🎉 Pipeline Complete! File is live at: {drive_link}")