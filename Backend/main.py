from pathlib import Path
from fastapi import FastAPI, HTTPException, BackgroundTasks
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel, Field
from datetime import datetime
from typing import Any, Optional
import json
import os
import queue
import threading
import asyncio
from concurrent.futures import ThreadPoolExecutor
from dotenv import load_dotenv
from fastapi.responses import StreamingResponse
import tempfile

# PPT Generator
from pptx_fill import run_ppt_job
from sync_ads_to_db_GA4 import sync_ga4_data

# --- EXCEL & GOOGLE ADS IMPORTS ---
from google.ads.googleads.client import GoogleAdsClient
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from drive_utils import get_drive_service, download_file, update_file

BASE_DIR = Path(__file__).resolve().parent
load_dotenv(BASE_DIR / ".env")

CLIENT_CONFIG_TABLE = os.environ.get("SUPABASE_CLIENT_TABLE", "google_ads_accounts")

app = FastAPI(title="Wheeler Automation API")

# CORS
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

executor = ThreadPoolExecutor(max_workers=5)
_sb = None

def get_supabase():
    global _sb
    if _sb is False:
        return None
    if _sb is not None:
        return _sb
    url = os.environ.get("SUPABASE_URL")
    key = os.environ.get("SUPABASE_KEY")
    if not url or not key:
        _sb = False
        return None
    from supabase import create_client
    _sb = create_client(url, key)
    return _sb


# ==========================================
# --- 1. PYDANTIC MODELS ---
# ==========================================

class GenerateRequest(BaseModel):
    client_id: str
    start_date: str
    end_date: str
    generate_ppt: bool = True
    generate_excel: bool = True

class ReportRequest(BaseModel):
    customer_id: str
    start_date: str
    end_date: str
    month_label: str

class AutomationLogCreate(BaseModel):
    status: str
    job_type: str = "report_generation"
    client_key: Optional[str] = None
    client_name: Optional[str] = None
    message: str
    duration_ms: Optional[int] = None
    triggered_by: str = "user"

class GeneratedReportCreate(BaseModel):
    folder_date: str
    report_range_start: str
    report_range_end: str
    client_key: str
    client_name: Optional[str] = None
    export_mode: str
    files: list[dict[str, Any]] = Field(default_factory=list)

class GeneratedReportFilesPatch(BaseModel):
    files: list[dict[str, Any]] = Field(default_factory=list)


class Ga4SyncStreamRequest(BaseModel):
    """Optional client_id = sync only that dealer; omit or null = all rows (CLI-style)."""

    client_id: Optional[str] = None


# ==========================================
# --- 2. EXCEL & GOOGLE ADS LOGIC ---
# ==========================================

HEADER_FILL = PatternFill(fill_type='solid', fgColor='1F4E79')
HEADER_FONT = Font(name='Calibri', color='FFFFFF', bold=True, size=11)
THIN_BORDER = Border(bottom=Side(border_style='thin', color='DDDDDD'))

def get_ads_client():
    yaml_path = os.environ.get('GOOGLE_ADS_YAML', str(BASE_DIR / 'Backend/google-ads.yaml'))
    
    if not os.path.exists(yaml_path):
        raise FileNotFoundError(f"CRITICAL: Could not find Google Ads config at {yaml_path}. Make sure 'google-ads.yaml' is in the same folder as main.py!")
        
    return GoogleAdsClient.load_from_storage(yaml_path)

def pull_auction_insights(ads_client, customer_id, campaign_id, start_date, end_date):
    query = f"""
        SELECT
            auction_insight.domain,
            metrics.auction_insight_search_impression_share,
            metrics.auction_insight_search_overlap_rate,
            metrics.auction_insight_search_outranking_share,
            metrics.auction_insight_search_position_above_rate,
            metrics.auction_insight_search_top_impression_percentage
        FROM campaign
        WHERE campaign.id = {campaign_id}
            AND segments.date BETWEEN '{start_date}' AND '{end_date}'
        ORDER BY metrics.auction_insight_search_impression_share DESC
    """
    ga_service = ads_client.get_service('GoogleAdsService')
    response = ga_service.search(customer_id=str(customer_id), query=query)
    rows = []
    for row in response:
        m = row.metrics
        rows.append([
            row.auction_insight.domain,
            f"{m.auction_insight_search_impression_share * 100:.1f}%",
            f"{m.auction_insight_search_overlap_rate * 100:.1f}%",
            f"{m.auction_insight_search_outranking_share * 100:.1f}%",
            f"{m.auction_insight_search_position_above_rate * 100:.1f}%",
            f"{m.auction_insight_search_top_impression_percentage * 100:.1f}%",
        ])
    return rows

def pull_search_terms(ads_client, customer_id, campaign_id, start_date, end_date):
    query = f"""
        SELECT
            search_term_view.search_term,
            segments.keyword.info.match_type,
            metrics.impressions,
            metrics.clicks,
            metrics.ctr,
            metrics.average_cpc,
            metrics.conversions
        FROM search_term_view
        WHERE campaign.id = {campaign_id}
            AND segments.date BETWEEN '{start_date}' AND '{end_date}'
        ORDER BY metrics.impressions DESC
        LIMIT 200
    """
    ga_service = ads_client.get_service('GoogleAdsService')
    response = ga_service.search(customer_id=str(customer_id), query=query)
    rows = []
    for row in response:
        m = row.metrics
        rows.append([
            row.search_term_view.search_term,
            row.segments.keyword.info.match_type.name.replace('_', ' ').title(),
            m.impressions,
            m.clicks,
            f"{m.ctr * 100:.2f}%",
            f"${m.average_cpc / 1_000_000:.2f}",
            round(m.conversions, 1),
        ])
    return rows

def write_excel_tab(workbook, tab_name, headers, rows, header_color='1F4E79'):
    if tab_name in workbook.sheetnames:
        del workbook[tab_name]
    
    ws = workbook.create_sheet(title=tab_name)
    header_fill = PatternFill(fill_type='solid', fgColor=header_color)
    
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal='center', vertical='center')
        
    ws.row_dimensions[1].height = 20
    
    for row_idx, row_data in enumerate(rows, 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = Alignment(vertical='center')
            cell.border = THIN_BORDER
            
    for col in ws.columns:
        max_length = 0
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except: pass
        ws.column_dimensions[col[0].column_letter].width = min(max_length + 4, 50)
    return ws

def run_auction_job(cfg, start_date, end_date, month_label):
    drive = get_drive_service()
    ads_client = get_ads_client()
    customer_id = str(cfg['google_ads_customer_id']).replace('-', '')

    AUCTION_HEADERS = ['Competitor Domain', 'Impression Share', 'Overlap Rate', 'Outranking Share', 'Position Above Rate', 'Top of Page Rate']
    SEARCH_HEADERS = ['Search Term', 'Match Type', 'Impressions', 'Clicks', 'CTR', 'Avg CPC', 'Conversions']

    with tempfile.TemporaryDirectory() as tmp_dir:
        local_excel = os.path.join(tmp_dir, 'auction_insights.xlsx')
        
        print("Downloading existing Excel from Google Drive...")
        download_file(drive, cfg['excel_drive_id'], local_excel)
        wb = load_workbook(local_excel)
        
        for campaign_id, campaign_name in cfg.get('campaigns', {}).items():
            print(f"Processing campaign: {campaign_name} ({campaign_id})")
            
            # Auction Insights
            auction_rows = pull_auction_insights(ads_client, customer_id, campaign_id, start_date, end_date)
            auction_tab = f"{month_label} - {campaign_name} - Auction"[:31]
            write_excel_tab(wb, auction_tab, AUCTION_HEADERS, auction_rows, '1F4E79')

            # Search Terms
            search_rows = pull_search_terms(ads_client, customer_id, campaign_id, start_date, end_date)
            search_tab = f"{month_label} - {campaign_name} - Search Terms"[:28] + "..." if len(f"{month_label} - {campaign_name} - Search Terms") > 31 else f"{month_label} - {campaign_name} - Search Terms"
            write_excel_tab(wb, search_tab, SEARCH_HEADERS, search_rows, '375623')

        wb.save(local_excel)
        wb.close() 
        print("Uploading updated Excel back to Google Drive...")
        updated = update_file(drive, cfg['excel_drive_id'], local_excel, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        
        return {
            'status': 'success',
            'drive_url': updated.get('webViewLink'),
            'message': f"Excel updated successfully with {len(cfg.get('campaigns', {}))} campaigns."
        }

# ==========================================
# --- 3. HELPER FUNCTIONS & WORKERS ---
# ==========================================

def _automation_log_to_api(row: dict) -> dict:
    return {
        "id": str(row.get("id", "")),
        "occurredAt": row.get("occurred_at"),
        "status": row.get("status"),
        "jobType": row.get("job_type"),
        "clientKey": row.get("client_key"),
        "clientName": row.get("client_name"),
        "message": row.get("message"),
        "durationMs": row.get("duration_ms"),
        "triggeredBy": row.get("triggered_by"),
    }

def _generated_report_to_api(row: dict) -> dict:
    files = row.get("files") or []
    if isinstance(files, str):
        try: files = json.loads(files)
        except: files = []
    fd = row.get("folder_date")
    rs = row.get("report_range_start")
    re = row.get("report_range_end")
    ca = row.get("created_at")
    return {
        "id": str(row.get("id", "")),
        "folderDate": fd.isoformat()[:10] if hasattr(fd, "isoformat") else str(fd)[:10],
        "reportRangeStart": rs.isoformat()[:10] if hasattr(rs, "isoformat") else str(rs)[:10],
        "reportRangeEnd": re.isoformat()[:10] if hasattr(re, "isoformat") else str(re)[:10],
        "createdAt": ca if isinstance(ca, str) else (ca.isoformat() if hasattr(ca, "isoformat") else str(ca)),
        "clientKey": row.get("client_key"),
        "clientName": row.get("client_name"),
        "exportMode": row.get("export_mode"),
        "files": files if isinstance(files, list) else [],
    }

def load_client_config(client_id: str):
    sb = get_supabase()
    if not sb:
        raise HTTPException(status_code=503, detail="Supabase not configured.")
    res = sb.table(CLIENT_CONFIG_TABLE).select("*").eq("client_id", client_id).execute()
    if not res.data:
        raise HTTPException(status_code=404, detail=f"Client '{client_id}' not found.")
    client_data = res.data[0]
    
    return {
        "client_name": client_data.get("descriptive_name", client_id),
        "customer_id": client_data.get("customer_id"),
        "template_drive_id": client_data.get("template_drive_id"),
        "output_file_drive_id": client_data.get("output_file_drive_id"),
        # THE FIX: Add the new Excel column to the config dictionary
        "excel_drive_id": client_data.get("excel_drive_id"), 
        "ga4_property_id": client_data.get("ga4_property_id"),
    }

def make_month_label(end_date: str) -> str:
    d = datetime.strptime(end_date, "%Y-%m-%d")
    return d.strftime("%B_%Y")

def _generate_worker_sync(request: GenerateRequest, cfg: dict, month_label: str, out_q: queue.Queue):
    result: dict = {}
    def forward(ev: dict):
        out_q.put({"event": "progress", "payload": ev})

    try:
        if request.generate_ppt:
            try:
                ppt = run_ppt_job(cfg, request.start_date, request.end_date, month_label, progress_cb=forward)
                result["ppt"] = ppt
            except Exception as e:
                result["ppt_error"] = str(e)
                forward({"kind": "log", "message": f"ERROR (PowerPoint): {e}"})

        if request.generate_excel:
            forward({"kind": "log", "message": "Starting Auction Insights Excel step…"})
            try:
                # Calls the internal function directly now!
                excel = run_auction_job(cfg, request.start_date, request.end_date, month_label)
                result["excel"] = excel
                msg = excel.get("message") if isinstance(excel, dict) else str(excel)
                forward({"kind": "log", "message": str(msg)})
            except Exception as e:
                result["excel_error"] = str(e)
                forward({"kind": "log", "message": f"ERROR (Excel): {e}"})

        out_q.put({"event": "result", "data": result})
    except Exception as e:
        out_q.put({"event": "error", "detail": str(e)})
    finally:
        out_q.put(None)


def _ga4_sync_worker(out_q: queue.Queue, client_id: Optional[str] = None):
    """Runs GA4 → Supabase `ga4_metrics` (sync_ads_to_db_GA4); streams log lines like generate-stream."""

    def forward(msg: str):
        out_q.put({"event": "progress", "payload": {"kind": "log", "message": msg}})

    try:
        summary = sync_ga4_data(log_callback=forward, client_id=client_id)
        out_q.put({"event": "result", "data": summary})
    except Exception as e:
        out_q.put({"event": "error", "detail": str(e)})
    finally:
        out_q.put(None)


def generate_client_auction_report(customer_id: str, start_date: str, end_date: str, month_label: str):
    sb = get_supabase()
    
    # THE FIX: Select the new excel_drive_id column
    client_res = sb.table(CLIENT_CONFIG_TABLE).select('customer_id, excel_drive_id').eq('customer_id', customer_id).execute()
    if not client_res.data:
        print(f"Error: Client {customer_id} not found in DB.")
        return
        
    client_data = client_res.data[0]
    
    # Query client campaigns
    try:
        campaigns_res = sb.table('client_campaigns').select('campaign_id, campaign_name').eq('customer_id', customer_id).execute()
        campaign_dict = {camp['campaign_id']: camp['campaign_name'] for camp in campaigns_res.data}
    except Exception:
        print("Warning: Could not fetch campaigns. Defaulting to empty.")
        campaign_dict = {}

    cfg = {
        'google_ads_customer_id': client_data['customer_id'],
        # THE FIX: Map the config to use the new Excel column!
        'excel_drive_id': client_data.get('excel_drive_id'), 
        'campaigns': campaign_dict
    }

    try:
        print(f"\n--- Starting Standalone Auction Job for {customer_id} ---")
        result = run_auction_job(cfg, start_date, end_date, month_label)
        print(f" -> DEBUG RESULT: {result}")
        print("--- Standalone Auction Job Complete! ---\n")
    except Exception as e:
        print(f"--- Auction Job Failed: {e} ---")


# ==========================================
# --- 4. API ENDPOINTS ---
# ==========================================

@app.get("/")
def root():
    return {"status": "Wheeler Automation API is running"}

@app.get("/clients")
def list_clients():
    sb = get_supabase()
    if not sb: return []
    try:
        res = sb.table(CLIENT_CONFIG_TABLE).select("*").execute()
        out = []
        for row in res.data or []:
            cid = row.get("client_id")
            if not cid: continue
            has_config = bool(row.get("template_drive_id") and row.get("output_file_drive_id"))
            out.append({
                "id": cid,
                "name": row.get("descriptive_name") or cid,
                "customer_id": row.get("customer_id"),
                "has_config": has_config,
            })
        return sorted(out, key=lambda x: str(x["name"]).lower())
    except Exception as e:
        print(f"Error fetching clients: {e}")
        return []

@app.post("/generate")
async def generate(request: GenerateRequest):
    try:
        s = datetime.strptime(request.start_date, "%Y-%m-%d")
        e = datetime.strptime(request.end_date, "%Y-%m-%d")
        if s >= e: raise HTTPException(status_code=400, detail="Start date must be before end date")
    except ValueError:
        raise HTTPException(status_code=400, detail="Invalid date format. Use YYYY-MM-DD")

    cfg = load_client_config(request.client_id)
    month_label = make_month_label(request.end_date)
    result = {}
    loop = asyncio.get_running_loop()

    if request.generate_ppt:
        try: result["ppt"] = await loop.run_in_executor(executor, run_ppt_job, cfg, request.start_date, request.end_date, month_label)
        except Exception as e: result["ppt_error"] = str(e)

    if request.generate_excel:
        try: result["excel"] = await loop.run_in_executor(executor, run_auction_job, cfg, request.start_date, request.end_date, month_label)
        except Exception as e: result["excel_error"] = str(e)

    return result

@app.post("/generate-stream")
async def generate_stream(request: GenerateRequest):
    try:
        s = datetime.strptime(request.start_date, "%Y-%m-%d")
        e = datetime.strptime(request.end_date, "%Y-%m-%d")
        if s >= e: raise HTTPException(status_code=400, detail="Start date must be before end date")
    except ValueError:
        raise HTTPException(status_code=400, detail="Invalid date format. Use YYYY-MM-DD")

    cfg = load_client_config(request.client_id)
    month_label = make_month_label(request.end_date)
    out_q: queue.Queue = queue.Queue()
    loop = asyncio.get_running_loop()

    threading.Thread(target=_generate_worker_sync, args=(request, cfg, month_label, out_q), daemon=True).start()

    async def event_iter():
        while True:
            item = await loop.run_in_executor(None, out_q.get)
            if item is None: break
            yield f"data: {json.dumps(item, default=str)}\n\n"
            await asyncio.sleep(0)

    headers = {"Cache-Control": "no-cache", "Connection": "keep-alive", "X-Accel-Buffering": "no"}
    return StreamingResponse(event_iter(), media_type="text/event-stream", headers=headers)


@app.post("/ga4-sync-stream")
async def ga4_sync_stream(body: Ga4SyncStreamRequest):
    """Stream GA4 channel + button metrics into `ga4_metrics` (sync_ads_to_db_GA4). client_id = one dealer; omit = all."""
    out_q: queue.Queue = queue.Queue()
    loop = asyncio.get_running_loop()
    cid = body.client_id.strip() if body.client_id and str(body.client_id).strip() else None
    threading.Thread(target=_ga4_sync_worker, args=(out_q, cid), daemon=True).start()

    async def event_iter():
        while True:
            item = await loop.run_in_executor(None, out_q.get)
            if item is None:
                break
            yield f"data: {json.dumps(item, default=str)}\n\n"
            await asyncio.sleep(0)

    headers = {"Cache-Control": "no-cache", "Connection": "keep-alive", "X-Accel-Buffering": "no"}
    return StreamingResponse(event_iter(), media_type="text/event-stream", headers=headers)


@app.post("/api/reports/generate-auction-insights")
async def trigger_auction_report(req: ReportRequest, background_tasks: BackgroundTasks):
    # Triggers the newly integrated function in the background!
    background_tasks.add_task(
        generate_client_auction_report, 
        req.customer_id, 
        req.start_date, 
        req.end_date, 
        req.month_label
    )
    return {"message": "Auction report generation started in the background."}

# ==========================================
# --- 5. LOGS & DB REPORTS ---
# ==========================================

@app.get("/automation-logs")
def list_automation_logs():
    sb = get_supabase()
    if not sb: raise HTTPException(status_code=503, detail="Supabase not configured.")
    try:
        res = sb.table("automation_logs").select("*").order("occurred_at", desc=True).limit(500).execute()
        return [_automation_log_to_api(r) for r in (res.data or [])]
    except Exception as e: raise HTTPException(status_code=503, detail=str(e))

@app.post("/automation-logs")
def create_automation_log(entry: AutomationLogCreate):
    sb = get_supabase()
    if not sb: raise HTTPException(status_code=503, detail="Supabase not configured.")
    try:
        row = {"status": entry.status, "job_type": entry.job_type, "client_key": entry.client_key, "client_name": entry.client_name, "message": entry.message, "duration_ms": entry.duration_ms, "triggered_by": entry.triggered_by}
        res = sb.table("automation_logs").insert(row).execute()
        if res.data: return _automation_log_to_api(res.data[0])
        return {"ok": True}
    except Exception as e: raise HTTPException(status_code=503, detail=str(e))

@app.get("/drive-reports")
def list_drive_reports():
    sb = get_supabase()
    if not sb: raise HTTPException(status_code=503, detail="Supabase not configured.")
    try:
        res = sb.table("generated_reports").select("*").order("created_at", desc=True).limit(500).execute()
        return [_generated_report_to_api(r) for r in (res.data or [])]
    except Exception as e: raise HTTPException(status_code=503, detail=str(e))

@app.post("/drive-reports")
def create_drive_report(entry: GeneratedReportCreate):
    sb = get_supabase()
    if not sb: raise HTTPException(status_code=503, detail="Supabase not configured.")
    try:
        row = {"folder_date": entry.folder_date[:10], "report_range_start": entry.report_range_start[:10], "report_range_end": entry.report_range_end[:10], "client_key": entry.client_key, "client_name": entry.client_name, "export_mode": entry.export_mode, "files": entry.files}
        res = sb.table("generated_reports").insert(row).execute()
        if res.data: return _generated_report_to_api(res.data[0])
        return {"ok": True}
    except Exception as e: raise HTTPException(status_code=503, detail=str(e))

@app.delete("/drive-reports/{report_id}")
def delete_drive_report(report_id: str):
    sb = get_supabase()
    if not sb: raise HTTPException(status_code=503, detail="Supabase not configured.")
    try:
        sb.table("generated_reports").delete().eq("id", str(report_id).strip()).execute()
        return {"ok": True, "deleted": True}
    except Exception as e: raise HTTPException(status_code=503, detail=str(e))

@app.patch("/drive-reports/{report_id}")
def patch_drive_report_files(report_id: str, body: GeneratedReportFilesPatch):
    sb = get_supabase()
    if not sb: raise HTTPException(status_code=503, detail="Supabase not configured.")
    rid = str(report_id).strip()
    try:
        if not body.files:
            sb.table("generated_reports").delete().eq("id", rid).execute()
            return {"ok": True, "deleted": True}
        res = sb.table("generated_reports").update({"files": body.files}).eq("id", rid).execute()
        if res.data: return _generated_report_to_api(res.data[0])
        return {"ok": True}
    except Exception as e: raise HTTPException(status_code=503, detail=str(e))

if __name__ == "__main__":
    # ==========================================
    # --- TEST MODE (Run Script Directly) ---
    # ==========================================
    # Change this to True when you want to test the script without starting the server
    RUN_TEST_MODE = True  
    
    if RUN_TEST_MODE:
        print("\n=== RUNNING IN TEST MODE ===")
        TEST_CUSTOMER_ID = "5691491477"
        START_DATE = "2026-03-01"
        END_DATE = "2026-03-31"
        MONTH_LABEL = "March 2026"
        
        # This will run the entire DB lookup, Google Ads fetch, and Excel upload instantly
        generate_client_auction_report(TEST_CUSTOMER_ID, START_DATE, END_DATE, MONTH_LABEL)
        print("=== TEST COMPLETE ===\n")
        
    # ==========================================
    # --- SERVER MODE (Run FastAPI) ---
    # ==========================================
    else:
        import uvicorn
        uvicorn.run(app, host="0.0.0.0", port=8000, reload=True)