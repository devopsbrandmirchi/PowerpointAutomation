import os
import tempfile
from pathlib import Path
from dotenv import load_dotenv
from supabase import create_client
from openpyxl import load_workbook
from datetime import datetime
import gc

from drive_utils import get_drive_service, download_file

BASE_DIR = Path(__file__).resolve().parent
load_dotenv(BASE_DIR / ".env")

# Your exact new Google Sheet ID
SHEET_ID = "18gkxaGU1TsWFEQAkvjIcrM8WKEIqcNVGzCsAk1cHPYM"   

# Fallback if Excel sheet is missing the "month" column
MANUAL_FALLBACK_MONTH_TEXT = "January 2026"

def get_supabase():
    url = os.environ.get("SUPABASE_URL")
    key = os.environ.get("SUPABASE_KEY")
    return create_client(url, key)

def clean_value(val):
    if not val or val == " --" or val == "--" or val == "-":
        return None
    val_str = str(val).replace('%', '').replace(',', '').strip()
    if '<' in val_str:
        return 0.09 # < 10% becomes 9%
    if '>' in val_str:
        return 0.90 # > 90% becomes 90%
    try:
        if '%' in str(val):
            return round(float(val_str) / 100.0, 4)
        return float(val_str)
    except ValueError:
        return str(val).strip()

def parse_month_to_date(month_text: str) -> str:
    """Converts 'January 2026' -> '2026-01-01'"""
    try:
        dt = datetime.strptime(str(month_text).strip(), "%B %Y")
        return dt.strftime("%Y-%m-%d")
    except Exception:
        try:
            dt = datetime.strptime(str(month_text).strip(), "%Y-%m-%d")
            return dt.strftime("%Y-%m-%d")
        except Exception:
            fallback = datetime.utcnow().strftime("%Y-%m-%d")
            return fallback

def get_account_mapping(sb):
    try:
        res = sb.table("google_ads_accounts").select("customer_id, client_id, descriptive_name").execute()
        mapping = {}
        if res.data:
            for row in res.data:
                cid = str(row.get("customer_id")).strip()
                mapping[cid] = {
                    "client_id": row.get("client_id"),
                    "account_name": row.get("descriptive_name")
                }
        return mapping
    except Exception as e:
        print(f"⚠️ Could not fetch account mapping: {e}")
        return {}

def sync_auction_insights():
    sb = get_supabase()
    drive = get_drive_service()
    
    print("=== STARTING AUCTION INSIGHTS SYNC ===")

    account_map = get_account_mapping(sb)

    with tempfile.TemporaryDirectory() as tmp_dir:
        sheet_path = os.path.join(tmp_dir, 'auction_data.xlsx')
        
        print("1. Downloading Sheet...")
        download_file(drive, SHEET_ID, sheet_path)
        
        print("2. Parsing Data...")
        wb = load_workbook(sheet_path, data_only=True)
        sheet = wb.active
        
        header_map = {}
        data_started = False
        db_rows = []
        
        for row in sheet.iter_rows(values_only=True):
            if not any(row):
                continue
                
            row_strs = [str(cell).lower().strip() if cell else "" for cell in row]
            
            # Map headers dynamically
            if not data_started:
                if "customer id" in row_strs:
                    for idx, col_name in enumerate(row_strs):
                        header_map[col_name] = idx
                    data_started = True
                continue
                
            if data_started:
                def get_val(aliases):
                    for alias in aliases:
                        for h_name, h_idx in header_map.items():
                            if alias in h_name:
                                return row[h_idx] if h_idx < len(row) else None
                    return None

                raw_cid = get_val(["customer id"])
                if not raw_cid or str(raw_cid).lower() == "total":
                    continue
                    
                cid = str(raw_cid).replace("-", "").strip()
                campaign = str(get_val(["campaign"]) or "Unknown").strip()
                domain = str(get_val(["display url domain", "display url"]) or "Unknown").strip()
                
                if not cid or not domain or domain.lower() == "none":
                    continue
                    
                mapped_data = account_map.get(cid, {})
                db_client_id = mapped_data.get("client_id")
                final_account_name = mapped_data.get("account_name") or str(get_val(["account"]) or "").strip()
                
                raw_month_text = get_val(["month"]) or MANUAL_FALLBACK_MONTH_TEXT
                report_date = parse_month_to_date(raw_month_text)

                impr_share = clean_value(get_val(["impr. share", "impr share"]))
                top_rate = clean_value(get_val(["top of page"]))
                abs_top_rate = clean_value(get_val(["abs. top"]))
                
                # ✅ NEW: Extracting Position Above Rate
                position_above_rate = clean_value(get_val(["position above"]))

                db_rows.append({
                    "customer_id":               cid,
                    "client_id":                 db_client_id,
                    "account_name":              final_account_name,
                    "report_date":               report_date,       
                    "campaign":                  campaign,
                    "domain":                    domain,
                    "impression_share":          impr_share,
                    "top_of_page_rate":          top_rate,
                    "absolute_top_of_page_rate": abs_top_rate,
                    "position_above_rate":       position_above_rate # ✅ Added to Supabase payload
                })
        
        wb.close()

        if not db_rows:
            print("❌ No valid data found to push.")
            return

        print(f"3. Deduplicating and formatting {len(db_rows)} rows...")
        unique_rows = {}
        for r in db_rows:
            key = (r["customer_id"], r["report_date"], r["domain"], r["campaign"])
            # If a duplicate exists, keep the one with the highest impression share
            if key in unique_rows:
                if r["impression_share"] and (not unique_rows[key]["impression_share"] or r["impression_share"] > unique_rows[key]["impression_share"]):
                    unique_rows[key] = r
            else:
                unique_rows[key] = r

        final_push_data = list(unique_rows.values())

        print(f"4. Pushing {len(final_push_data)} clean rows to Supabase...")
        
        batch_size = 1000
        for i in range(0, len(final_push_data), batch_size):
            chunk = final_push_data[i:i + batch_size]
            sb.table("google_ads_auction_master").upsert(
                chunk,
                on_conflict="customer_id,report_date,domain,campaign"  
            ).execute()
            
        print("=== 🚀 INGEST COMPLETE ===")
        gc.collect()

if __name__ == "__main__":
    sync_auction_insights()